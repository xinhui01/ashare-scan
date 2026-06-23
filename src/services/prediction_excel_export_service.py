"""Excel export for limit-up prediction results."""
from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Any, Dict, Iterable, List, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from src.services.market_focus_advice_service import (
    format_market_focus_advice_lines,
    resolve_market_focus_advice,
)


CandidateSpec = Tuple[str, str, List[Tuple[str, str]]]
UNCONFIRMED_TEXT = "未确认"
UNCONFIRMED_AUCTION_TEXT = "需9:25后竞价确认"


CANDIDATE_SPECS: List[CandidateSpec] = [
    (
        "保留涨停",
        "continuation_candidates",
        [
            ("code", "代码"), ("name", "名称"), ("industry", "行业"), ("theme", "题材"),
            ("consecutive_boards", "连板数"), ("change_pct", "涨幅%"),
            ("first_board_time", "首封时间"), ("accumulation_score", "潜伏分"),
            ("relative_strength_score", "强弱分"), ("score", "预测分"),
            ("_confirm_text", "确认"), ("_auction_text", "竞价/开盘"),
            ("_result_text", "结果"), ("reasons", "预测依据"),
        ],
    ),
    (
        "二波接力",
        "first_board_candidates",
        [
            ("code", "代码"), ("name", "名称"), ("industry", "行业"), ("theme", "题材"),
            ("change_pct", "今日涨幅%"), ("volume_ratio", "爆量倍数"),
            ("dist_ma5_pct", "距MA5%"), ("accumulation_score", "潜伏分"),
            ("relative_strength_score", "强弱分"), ("score", "预测分"),
            ("_confirm_text", "确认"), ("_auction_text", "竞价/开盘"),
            ("_result_text", "结果"), ("reasons", "预测依据"),
        ],
    ),
    (
        "首板涨停",
        "fresh_first_board_candidates",
        [
            ("code", "代码"), ("name", "名称"), ("industry", "行业"), ("theme", "题材"),
            ("change_pct", "今日涨幅%"), ("volume_ratio", "量比"),
            ("dist_ma5_pct", "距MA5%"), ("trend_5d", "5日涨幅%"),
            ("accumulation_score", "潜伏分"), ("relative_strength_score", "强弱分"),
            ("score", "预测分"),
            ("_confirm_text", "确认"), ("_auction_text", "竞价/开盘"),
            ("_result_text", "结果"), ("reasons", "预测依据"),
        ],
    ),
    (
        "反包",
        "broken_board_wrap_candidates",
        [
            ("code", "代码"), ("name", "名称"), ("industry", "行业"), ("theme", "题材"),
            ("predict_type", "形态"), ("change_pct", "今日涨幅%"),
            ("prior_lu_date", "前涨停日"), ("wrap_gap_pct", "反包缺口%"),
            ("days_since_lu", "距前涨停"), ("accumulation_score", "潜伏分"),
            ("relative_strength_score", "强弱分"), ("score", "预测分"),
            ("_confirm_text", "确认"), ("_auction_text", "竞价/开盘"),
            ("_result_text", "结果"), ("reasons", "预测依据"),
        ],
    ),
    (
        "趋势涨停",
        "trend_limit_up_candidates",
        [
            ("code", "代码"), ("name", "名称"), ("industry", "行业"), ("theme", "题材"),
            ("change_pct", "今日涨幅%"), ("ma_spread_pct", "均线差%"),
            ("ma20_slope_pct", "MA20斜率%"), ("trend_5d", "5日涨幅%"),
            ("volume_ratio", "量比"), ("accumulation_score", "潜伏分"),
            ("relative_strength_score", "强弱分"), ("score", "预测分"),
            ("_confirm_text", "确认"), ("_auction_text", "竞价/开盘"),
            ("_result_text", "结果"), ("reasons", "预测依据"),
        ],
    ),
]


def _safe_sheet_name(name: str) -> str:
    return str(name or "Sheet")[:31]


def _record_value(record: Dict[str, Any], key: str) -> Any:
    value = record.get(key)
    if value is None:
        return ""
    if isinstance(value, float):
        return round(value, 4)
    return value


def _safe_float(value: Any) -> float | None:
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _iter_candidate_records(prediction: Dict[str, Any]) -> Iterable[Dict[str, Any]]:
    for _sheet_name, key, _columns in CANDIDATE_SPECS:
        for rec in prediction.get(key) or []:
            if isinstance(rec, dict):
                yield rec


def _build_operation_hint(prediction: Dict[str, Any]) -> str:
    ctx = prediction.get("compare_context") or {}
    hints: List[str] = []
    candidates = list(_iter_candidate_records(prediction))
    if candidates and not any((rec.get("opening_confirmation") or {}) for rec in candidates):
        hints.append(
            "未做竞价/开盘确认：候选仅为观察池，不应直接按表买入；"
            "次日9:25后先执行竞价确认。"
        )
    avg_cont_rate = _safe_float(ctx.get("avg_continuation_rate"))
    if avg_cont_rate is not None and avg_cont_rate < 15:
        pair_count = int(ctx.get("pair_count") or 0)
        hints.append(
            f"近{pair_count}组平均晋级率仅{avg_cont_rate:.1f}%，接力环境偏冷；"
            "保留涨停需降低优先级，放弃无确认票。"
        )
    if candidates:
        hints.append("热门行业/题材只表示涨停聚集，不等同买入方向。")
    return "\n".join(hints)


def _theme_for_record(record: Dict[str, Any], compare_context: Dict[str, Any]) -> str:
    theme = str(record.get("theme") or record.get("theme_name") or "").strip()
    if theme:
        return theme
    code = str(record.get("code") or "").strip().zfill(6)
    code_theme_map = compare_context.get("code_theme_map") or {}
    return str(code_theme_map.get(code) or "")


def _prepare_records(records: Iterable[Dict[str, Any]], compare_context: Dict[str, Any]) -> List[Dict[str, Any]]:
    out: List[Dict[str, Any]] = []
    for raw in records or []:
        if not isinstance(raw, dict):
            continue
        rec = dict(raw)
        rec["theme"] = _theme_for_record(rec, compare_context)
        confirm = rec.get("opening_confirmation") or {}
        if isinstance(confirm, dict) and confirm:
            rec["_confirm_text"] = confirm.get("status") or ""
            rec["_auction_text"] = confirm.get("summary") or confirm.get("auction_status") or ""
        else:
            rec["_confirm_text"] = UNCONFIRMED_TEXT
            rec["_auction_text"] = UNCONFIRMED_AUCTION_TEXT
        rec["_result_text"] = rec.get("result") or rec.get("_result_text") or ""
        out.append(rec)
    return out


def _write_rows(ws, headers: List[str], rows: List[List[Any]]) -> None:
    ws.append(headers)
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in rows:
        ws.append(row)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    ws.freeze_panes = "A2"


def _autosize(ws, *, max_width: int = 42) -> None:
    for idx, column_cells in enumerate(ws.columns, start=1):
        width = 8
        for cell in column_cells:
            text = str(cell.value or "")
            if not text:
                continue
            width = max(width, min(max_width, max(len(part) for part in text.splitlines()) + 2))
        ws.column_dimensions[get_column_letter(idx)].width = width


def _write_summary(wb: Workbook, prediction: Dict[str, Any]) -> None:
    ws = wb.active
    ws.title = "汇总"
    ctx = prediction.get("compare_context") or {}
    specs_count = {name: len(prediction.get(key) or []) for name, key, _cols in CANDIDATE_SPECS}
    market_focus_advice = resolve_market_focus_advice(prediction)
    market_focus_lines = format_market_focus_advice_lines(market_focus_advice)
    rows = [
        ["交易日", prediction.get("trade_date", "")],
        ["回溯天数", prediction.get("lookback_days", "")],
        ["生成时间", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
        ["市场情绪分", ctx.get("sentiment_score", "")],
        ["基础情绪分", ctx.get("sentiment_base_score", "")],
        ["题材情绪增量", ctx.get("theme_sentiment_delta", "")],
        ["市场状态", ctx.get("market_state_label", "")],
        ["情绪打法", (ctx.get("market_state_strategy") or {}).get("label", "")],
        ["轮动分", (ctx.get("market_rotation") or {}).get("rotation_score", "")],
        ["行情打法建议", "\n".join(market_focus_lines)],
        ["今日重点池", market_focus_advice.get("focus_text", "") if market_focus_advice else ""],
        ["备选观察", market_focus_advice.get("secondary_text", "") if market_focus_advice else ""],
        ["谨慎/回避池", market_focus_advice.get("avoid_text", "") if market_focus_advice else ""],
        ["预测摘要", prediction.get("summary", "")],
    ]
    for name, count in specs_count.items():
        rows.append([f"{name}数量", count])
    operation_hint = _build_operation_hint(prediction)
    if operation_hint:
        rows.append(["操作提示", operation_hint])
    for row in rows:
        ws.append(row)
    for cell in ws["A"]:
        cell.font = Font(bold=True)
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    _autosize(ws, max_width=80)


def _write_candidate_sheet(
    wb: Workbook,
    prediction: Dict[str, Any],
    sheet_name: str,
    key: str,
    columns: List[Tuple[str, str]],
) -> None:
    ctx = prediction.get("compare_context") or {}
    records = _prepare_records(prediction.get(key) or [], ctx)
    ws = wb.create_sheet(_safe_sheet_name(sheet_name))
    headers = [label for _field, label in columns]
    rows = [[_record_value(rec, field) for field, _label in columns] for rec in records]
    _write_rows(ws, headers, rows)
    _autosize(ws)


def _write_theme_sheet(wb: Workbook, prediction: Dict[str, Any]) -> None:
    ws = wb.create_sheet("题材资金")
    ctx = prediction.get("compare_context") or {}
    fund_map = ctx.get("theme_fund_score_map") or {}
    acc_map = ctx.get("theme_fund_accumulation_map") or {}
    burst_map = ctx.get("theme_breakout_map") or {}
    groups = {
        str(g.get("name") or ""): g
        for g in ((prediction.get("theme_prediction") or {}).get("groups") or [])
        if isinstance(g, dict)
    }
    names = sorted(set(fund_map) | set(acc_map) | set(burst_map) | set(groups))
    headers = ["题材", "阶段", "机会分", "潜伏分", "资金分", "爆发分", "候选数", "角色分布"]
    rows: List[List[Any]] = []
    for name in names:
        group = groups.get(name) or {}
        counts = group.get("counts") or {}
        role_text = " / ".join(f"{k}:{v}" for k, v in counts.items() if v)
        rows.append([
            name,
            group.get("phase", ""),
            group.get("opportunity_score", ""),
            acc_map.get(name, ""),
            fund_map.get(name, ""),
            burst_map.get(name, ""),
            group.get("candidate_count", ""),
            role_text,
        ])
    _write_rows(ws, headers, rows)
    _autosize(ws)


def export_prediction_to_excel(prediction: Dict[str, Any], path: str | Path) -> Path:
    """Export a prediction payload to an .xlsx file and return the written path."""
    if not isinstance(prediction, dict) or not prediction:
        raise ValueError("没有可导出的预测结果")
    out_path = Path(path)
    if out_path.suffix.lower() != ".xlsx":
        out_path = out_path.with_suffix(".xlsx")
    out_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    _write_summary(wb, prediction)
    for sheet_name, key, columns in CANDIDATE_SPECS:
        _write_candidate_sheet(wb, prediction, sheet_name, key, columns)
    _write_theme_sheet(wb, prediction)
    wb.save(out_path)
    return out_path
