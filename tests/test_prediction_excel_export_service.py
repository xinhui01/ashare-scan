from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from src.services.prediction_excel_export_service import export_prediction_to_excel


def test_export_prediction_to_excel_writes_summary_and_candidate_sheets(tmp_path: Path):
    prediction = {
        "trade_date": "20260621",
        "lookback_days": 5,
        "summary": "测试摘要",
        "compare_context": {
            "sentiment_score": 72,
            "sentiment_base_score": 65,
            "theme_sentiment_delta": 7,
            "market_state_label": "轮动日",
            "market_state_strategy": {"label": "首板新题材 / 避开老主线"},
            "market_rotation": {"rotation_score": 42},
            "pair_count": 5,
            "avg_continuation_rate": 10.8,
            "theme_fund_score_map": {"机器人": 68},
            "theme_fund_accumulation_map": {"机器人": 18},
            "theme_breakout_map": {"机器人": 34},
        },
        "continuation_candidates": [
            {
                "code": "300001",
                "name": "机甲龙头",
                "industry": "通用设备",
                "theme": "机器人",
                "consecutive_boards": 3,
                "first_board_time": "09:35",
                "accumulation_score": 12,
                "score": 88,
                "change_pct": 10.0,
                "reasons": "3连板+30 / 题材资金潜18/爆34+6",
            }
        ],
        "first_board_candidates": [],
        "fresh_first_board_candidates": [],
        "broken_board_wrap_candidates": [],
        "trend_limit_up_candidates": [],
        "theme_prediction": {
            "groups": [
                {
                    "name": "机器人",
                    "source": "概念",
                    "phase": "主升",
                    "opportunity_score": 82,
                    "candidate_count": 1,
                    "counts": {"core": 1, "relay": 0, "repair": 0, "replenish": 0, "watch": 0},
                }
            ]
        },
    }
    out = tmp_path / "prediction.xlsx"

    export_prediction_to_excel(prediction, out)

    wb = load_workbook(out)
    assert wb.sheetnames == ["汇总", "保留涨停", "二波接力", "首板涨停", "反包", "趋势涨停", "题材资金"]
    assert wb["汇总"]["A1"].value == "交易日"
    assert wb["汇总"]["B1"].value == "20260621"
    headers = [cell.value for cell in wb["保留涨停"][1]]
    assert "潜伏分" in headers
    assert "预测依据" in headers
    assert wb["保留涨停"]["A2"].value == "300001"
    assert wb["保留涨停"]["K2"].value == "未确认"
    assert wb["保留涨停"]["L2"].value == "需9:25后竞价确认"
    summary_values = [cell.value for cell in wb["汇总"]["B"] if cell.value]
    assert "轮动日" in summary_values
    assert "首板新题材 / 避开老主线" in summary_values
    assert 42 in summary_values
    assert any("首板涨停(0只" in str(value) for value in summary_values)
    assert any("保留涨停/连板(1只)" in str(value) for value in summary_values)
    assert any("候选仅为观察池" in str(value) for value in summary_values)
    assert any("平均晋级率仅10.8%" in str(value) for value in summary_values)
    assert wb["题材资金"]["A2"].value == "机器人"
    assert wb["题材资金"]["E2"].value == 68
